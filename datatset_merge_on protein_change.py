import os
import pandas as pd
import re
import openpyxl
from openpyxl import Workbook, formatting
from openpyxl.styles import Font, NamedStyle, Alignment, PatternFill, Border, Side, Color
from openpyxl.worksheet import page
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.differential import DifferentialStyle

pd.set_option('display.max_columns', 30)
pd.set_option('display.expand_frame_repr', False)
pd.set_option('display.max_rows', None)

print("Call the first function, fm_tso, with the case number in quotes")

def fm_tso(case_number):
        fm = pd.read_excel("TSO500 cases to be cut.xlsx", "Sheet1").fillna("none")
        global file
        file = case_number + "_CombinedVariantOutput.xlsx"
        sheet = case_number + "_CombinedVariantOutput"
        global tso
        tso = pd.read_excel(file, sheet).fillna("none")

        #get MS stability and tmb before the aa nomenclature change
        casemetmb = fm.loc[fm["Case Number"] == case_number]
        global mstmb
        mstmb = []
        for i in casemetmb.iloc[0,:]:
                if str(i).startswith(("[MS", "[TMB")):
                        mstmb.append(i)
                else:
                        pass

        #change single letter codes in fm dataset to 3 letter codes
        fmcol = list(range(7, 129, 2))
        for f in fmcol:
                pdf = ''
                pdff =''
                pdf = []
                for i in fm.iloc[:,f]:
                        t = i.replace("G", "gly").replace("C", "cys").replace('D','asp').replace('S','ser').replace('Q','gln').replace('K','lys').replace('I','ile')\
                        .replace('P', 'pro').replace('T', 'thr').replace( 'F','phe').replace('N', 'asn').replace('H','his').replace('L', 'leu').replace('R', 'arg').replace('W', 'trp')\
                        .replace('A', 'ala').replace('V','val').replace('E','glu').replace('Y','tyr').replace('M','met').replace("*", "Ter")
                        s = t.replace("gly", "GLY").replace("cys", "CYS").replace('asp','ASP').replace('ser','SER').replace('gln','GLN').replace('lys','LYS').replace('ile','ILE')\
                        .replace('pro', 'PRO').replace('thr', 'THR').replace( 'phe','PHE').replace('asn', 'ASN').replace('his','HIS').replace('leu', 'LEU').replace('arg', 'ARG')\
                        .replace('trp', 'TRP').replace('ala', 'ALA').replace('val','VAL').replace('glu','GLU').replace('tyr','TYR').replace('met','MET')
                        l = s.replace("VALUSER", "VUS")
                        pdf.append(l)
                pdff = pd.DataFrame(pdf)
                fm.iloc[:,f] = pdff
        
        #change 3 letter codes in tso dataset to match fm dataset format
        tsodf = []
        for i in tso.iloc[:,7]:
                s = i.replace("Gly", "GLY").replace("Cys", "CYS").replace('Asp','ASP').replace('Ser','SER').replace('Gln','GLN').replace('Lys','LYS').replace('Ile','ILE')\
                        .replace('Pro', 'PRO').replace('Thr', 'THR').replace( 'Phe','PHE').replace('Asn', 'ASN').replace('His','HIS').replace('Leu', 'LEU').replace('Arg', 'ARG')\
                        .replace('Trp', 'TRP').replace('Ala', 'ALA').replace('Val','VAL').replace('Glu','GLU').replace('Tyr','TYR').replace('Met','MET')
                tsodf.append(s)
        tsodff = pd.DataFrame(tsodf)
        tso.iloc[:,7] = tsodff

        #merge columns of genes and protein changes in fm dataset
        genes = list(range(6,127,2))
        abn = list(range(7,128,2))

        for i,j in zip(genes, abn):
                fm.iloc[:,i] = fm.iloc[:,i] + "_" + fm.iloc[:,j]
        fm.drop(fm.columns[abn], axis=1, inplace=True)


        #GET CASE from curated (genes_abn) fm df in shell
        case = fm.loc[fm["Case Number"] == case_number]
        global caset
        caset = case.transpose()
        case_cull1 = caset.iloc[6:,0]

        #have series with just genes and no MS/TMB
        case_cull2 = []
        for i in case_cull1:
                if str(i).startswith(("Mic", "Tum")):
                        pass
                else:
                        case_cull2.append(i)

        #split the amp and loss genes away from df to append to fm_tso_info
        amp_loss = []
        case_cull2_2 = []
        for i in case_cull2:
                if "amplification" in i or "loss" in i:
                        amp_loss.append(i)
                else:
                        case_cull2_2.append(i)
        global amp_loss_2
        amp_loss_2 = pd.DataFrame(amp_loss)
        print(amp_loss_2)
        if amp_loss_2.empty:
                amp_loss_2 = ["none"]
                amp_loss_2 = pd.DataFrame(amp_loss_2)      
        
        #split for multiple variants in fm dataset
        split_var = []
        for i in case_cull2_2:
                if "," in i:
                        vus_match = re.search(r"(.*)[_][[](.*)[,]\s(.*)[]]", i)
                        gene = vus_match.group(1)
                        vus1 = vus_match.group(2)
                        vus2 = vus_match.group(3)
                        vus1_1 = gene + "_[" + vus1 +"]"
                        vus2_2 = gene + "_[" + vus2 +"]"
                        split_var.extend([vus1_1,vus2_2])
                else:
                        split_var.append(i)
        global split_varpd
        split_varpd = pd.DataFrame(split_var)
        
        print(split_varpd)
        print("Look at split_varpd and amp_loss_2 to manually change entries with more than one entry. ex...split_varpd.iloc[13,0] = ALK_[(VUS) LYS567ASP]")
        print("For next function fm_tso_combine, look at tso.iloc[30:55,0] to get row for start of genes (include row with column headings), tsop_start, and row for end of gene amps, tso_info_end")
        #add rows to df with df.loc[next index after last] = ....

        
def fm_tso_combine(split_varpd, tsop_start, tso_info_end):
        #create dataframes of tso abn and fm
        global fmp
        fmp=split_varpd
        fmp.columns = ["FM_Variants"]
        tsop = pd.DataFrame(tso.iloc[tsop_start:, [0,5,6,7,8,9]])
        tsop.reset_index(drop=True, inplace=True)

        #for TSO500, extract protein variant to then combine with gene
        abnp = []
        for i in tsop.iloc[:,3]:
                if i.startswith("N"):
                        tso_match = re.search(r"(.*)[(](.*)[)].*", i)
                        pro = tso_match.group(2)
                        abnp.append(pro)
                else:
                        abnp.append("none")
        abnpdf = pd.DataFrame(abnp)
        tsop["protein"] = abnpdf

        #get rid of none_none in fm dataframe
        global fmabnp
        fmabnp = []
        for i in fmp.iloc[:,0]:
                if i != "none_none":
                        fm_match = re.search(r"(.*)[_[](.*)[]].*", i)
                        pro = fm_match.group(2)
                        fmabnp.append(pro)
                else:
                        fmabnp.append("none")
          
        #get rid of vus
        fma_novus = []
        for i in fmabnp:
                fma_rvus = i.replace("(VUS) ", '')
                fma_novus.append(fma_rvus)
        fmabnpdf = pd.DataFrame(fma_novus)
        fmp["protein"] = fmabnpdf

        #get rid of rows with none_none
        fmp = fmp[~fmp.FM_Variants.str.contains("none_none")]


        #merge the dfs on the protein using the fm keys only!
        merge = fmp.merge(tsop, how="left")
        merge.columns = ["FM_variant", "Protein in both", "TSO_Gene", "Allele Frequency", "Depth", "P-Dot Notation", "C-Dot Notation", "Consequence"]


        tso500genes = ['ABL1', 'ABL2', 'ACVR1', 'ACVR1B', 'AKT1', 'AKT2', 'AKT3', 'ALK', 'ALOX12B', 'ANKRD11', 'ANKRD26', 'APC', 'AR', 'ARAF', 'ARFRP1', 'ARID1A',
                       'ARID1B', 'ARID2', 'ARID5B', 'ASXL1', 'ASXL2', 'ATM', 'ATR', 'ATRX', 'AURKA', 'AURKB', 'AXIN1', 'AXIN2', 'AXL', 'B2M', 'BAP1', 'BARD1', 'BBC3',
                       'BCL10', 'BCL2', 'BCL2L1', 'BCL2L11', 'BCL2L2', 'BCL6', 'BCOR', 'BCORL1', 'BCR', 'BIRC3', 'BLM', 'BMPR1A', 'BRAF', 'BRCA1', 'BRCA2', 'BRD4', 'BRIP1',
                       'BTG1', 'BTK', 'C11orf30', 'CALR', 'CARD11', 'CASP8', 'CBFB', 'CBL', 'CCND1', 'CCND2', 'CCND3', 'CCNE1', 'CD274', 'CD276', 'CD74', 'CD79A', 'CD79B',
                       'CDC73', 'CDH1', 'CDK12', 'CDK4', 'CDK6', 'CDK8', 'CDKN1A', 'CDKN1B', 'CDKN2A', 'CDKN2B', 'CDKN2C', 'CEBPA', 'CENPA', 'CHD2', 'CHD4', 'CHEK1', 'CHEK2', 'CIC', 'CREBBP',
                       'CRKL', 'CRLF2', 'CSF1R', 'CSF3R', 'CSNK1A1', 'CTCF', 'CTLA4', 'CTNNA1', 'CTNNB1', 'CUL3', 'CUX1', 'CXCR4', 'CYLD', 'DAXX', 'DCUN1D1', 'DDR2', 'DDX41', 'DHX15', 'DICER1',
                       'DIS3', 'DNAJB1', 'DNMT1', 'DNMT3A', 'DNMT3B', 'DOT1L', 'E2F3', 'EED', 'EGFL7', 'EGFR', 'EIF1AX', 'EIF4A2', 'EIF4E', 'EML4', 'EP300', 'EPCAM', 'EPHA3', 'EPHA5', 'EPHA7',
                       'EPHB1', 'ERBB2', 'ERBB3', 'ERBB4', 'ERCC1', 'ERCC2', 'ERCC3', 'ERCC4', 'ERCC5', 'ERG', 'ERRFI1', 'ESR1', 'ETS1', 'ETV1', 'ETV4', 'ETV5', 'ETV6', 'EWSR1', 'EZH2', 'FAM123B',
                       'FAM175A', 'FAM46C', 'FANCA', 'FANCC', 'FANCD2', 'FANCE', 'FANCF', 'FANCG', 'FANCI', 'FANCL', 'FAS', 'FAT1', 'FBXW7', 'FGF1', 'FGF10', 'FGF14', 'FGF19', 'FGF2', 'FGF23', 'FGF3',
                       'FGF4', 'FGF5', 'FGF6', 'FGF7', 'FGF8', 'FGF9', 'FGFR1', 'FGFR2', 'FGFR3', 'FGFR4', 'FH', 'FLCN', 'FLI1', 'FLT1', 'FLT3', 'FLT4', 'FOXA1', 'FOXL2', 'FOXO1', 'FOXP1', 'FRS2', 'FUBP1',
                       'FYN', 'GABRA6', 'GATA1', 'GATA2', 'GATA3', 'GATA4', 'GATA6', 'GEN1', 'GID4', 'GLI1', 'GNA11', 'GNA13', 'GNAQ', 'GNAS', 'GPR124', 'GPS2', 'GREM1', 'GRIN2A', 'GRM3', 'GSK3B', 'H3F3A',
                       'H3F3B', 'H3F3C', 'HGF', 'HIST1H1C', 'HIST1H2BD', 'HIST1H3A', 'HIST1H3B', 'HIST1H3C', 'HIST1H3D', 'HIST1H3E', 'HIST1H3F', 'HIST1H3G', 'HIST1H3H', 'HIST1H3I', 'HIST1H3J', 'HIST2H3A',
                       'HIST2H3C', 'HIST2H3D', 'HIST3H3', 'HLA-A', 'HLA-B', 'HLA-C', 'HNF1A', 'HNRNPK', 'HOXB13', 'HRAS', 'HSD3B1', 'HSP90AA1', 'ICOSLG', 'ID3', 'IDH1', 'IDH2', 'IFNGR1', 'IGF1', 'IGF1R', 'IGF2',
                       'IKBKE', 'IKZF1', 'IL10', 'IL7R', 'INHA', 'INHBA', 'INPP4A', 'INPP4B', 'INSR', 'IRF2', 'IRF4', 'IRS1', 'IRS2', 'JAK1', 'JAK2', 'JAK3', 'JUN', 'KAT6A', 'KDM5A', 'KDM5C', 'KDM6A', 'KDR', 'KEAP1',
                       'KEL', 'KIF5B', 'KIT', 'KLF4', 'KLHL6', 'KMT2B', 'KMT2C', 'KMT2D', 'KMT2A', 'KRAS', 'LAMP1', 'LATS1', 'LATS2', 'LMO1', 'LRP1B', 'LYN', 'LZTR1', 'MAGI2', 'MALT1', 'MAP2K1', 'MAP2K2', 'MAP2K4', 'MAP3K1',
                       'MAP3K13', 'MAP3K14', 'MAP3K4', 'MAPK1', 'MAPK3', 'MAX', 'MCL1', 'MDC1', 'MDM2', 'MDM4', 'MED12', 'MEF2B', 'MEN1', 'MET', 'MGA', 'MITF', 'MLH1', 'MLL', 'MLL2','MLLT3', 'MPL', 'MRE11A', 'MSH2', 'MSH3',
                       'MSH6', 'MST1', 'MST1R', 'MTOR', 'MUTYH', 'MYB', 'MYC', 'MYCL1', 'MYCN', 'MYD88', 'MYOD1', 'NAB2', 'NBN', 'NCOA3', 'NCOR1', 'NEGR1', 'NF1', 'NF2', 'NFE2L2', 'NFKBIA', 'NKX2-1', 'NKX3-1', 'NOTCH1',
                       'NOTCH2', 'NOTCH3', 'NOTCH4', 'NPM1', 'NRAS', 'NRG1', 'NSD1', 'NTRK1', 'NTRK2', 'NTRK3', 'NUP93', 'NUTM1', 'PAK1', 'PAK3', 'PAK7', 'PALB2', 'PARK2', 'PARP1', 'PAX3', 'PAX5', 'PAX7', 'PAX8', 'PBRM1',
                       'PDCD1', 'PDCD1LG2', 'PDGFRA', 'PDGFRB', 'PDK1', 'PDPK1', 'PGR', 'PHF6', 'PHOX2B', 'PIK3C2B', 'PIK3C2G', 'PIK3C3', 'PIK3CA', 'PIK3CB', 'PIK3CD', 'PIK3CG', 'PIK3R1', 'PIK3R2', 'PIK3R3', 'PIM1', 'PLCG2',
                       'PLK2', 'PMAIP1', 'PMS1', 'PMS2', 'PNRC1', 'POLD1', 'POLE', 'PPARG', 'PPM1D', 'PPP2R1A', 'PPP2R2A', 'PPP6C', 'PRDM1', 'PREX2', 'PRKAR1A', 'PRKCI', 'PRKDC', 'PRSS8', 'PTCH1', 'PTEN', 'PTPN11', 'PTPRD',
                       'PTPRS', 'PTPRT', 'QKI', 'RAB35', 'RAC1', 'RAD21', 'RAD50', 'RAD51', 'RAD51B', 'RAD51C', 'RAD51D', 'RAD52', 'RAD54L', 'RAF1', 'RANBP2', 'RARA', 'RASA1', 'RB1', 'RBM10', 'RECQL4', 'REL', 'RET', 'RFWD2',
                       'RHEB', 'RHOA', 'RICTOR', 'RIT1', 'RNF43', 'ROS1', 'RPS6KA4', 'RPS6KB1', 'RPS6KB2', 'RPTOR', 'RUNX1', 'RUNX1T1', 'RYBP', 'SDHA', 'SDHAF2', 'SDHB', 'SDHC', 'SDHD', 'SETBP1', 'SETD2', 'SF3B1', 'SH2B3',
                       'SH2D1A', 'SHQ1', 'SLIT2', 'SLX4', 'SMAD2', 'SMAD3', 'SMAD4', 'SMARCA4', 'SMARCB1', 'SMARCD1', 'SMC1A', 'SMC3', 'SMO', 'SNCAIP', 'SOCS1', 'SOX10', 'SOX17', 'SOX2', 'SOX9', 'SPEN', 'SPOP', 'SPTA1', 'SRC',
                       'SRSF2', 'STAG1', 'STAG2', 'STAT3', 'STAT4', 'STAT5A', 'STAT5B', 'STK11', 'STK40', 'SUFU', 'SUZ12', 'SYK', 'TAF1', 'TBX3', 'TCEB1', 'TCF3', 'TCF7L2', 'TERC', 'TERT', 'TET1', 'TET2', 'TFE3', 'TFRC', 'TGFBR1',
                       'TGFBR2', 'TMEM127', 'TMPRSS2', 'TNFAIP3', 'TNFRSF14', 'TOP1', 'TOP2A', 'TP53', 'TP63', 'TRAF2', 'TRAF7', 'TSC1', 'TSC2', 'TSHR', 'U2AF1', 'VEGFA', 'VHL', 'VTCN1', 'WISP3', 'WT1', 'XIAP', 'XPO1', 'XRCC2',
                       'YAP1', 'YES1', 'ZBTB2', 'ZBTB7A', 'ZFHX3', 'ZNF217', 'ZNF703', 'ZRSR2']


        #get genes from fm_variants for statistics, see if in tso, and see if not detected
        yesnogenes = []
        for g, p in zip(merge.iloc[:,0], merge.iloc[:,5]):
                gene_only = re.search(r"(.*)[_][[](.*)[]]", g)
                fmg = gene_only.group(1)
                if fmg in tso500genes and pd.isna(p):
                        yesnogenes.append("Not detected")
                else:
                        yesnogenes.append("")

        yesnogenes_df = pd.DataFrame(yesnogenes)
        merge["Gene covered by TSO500?"] = yesnogenes_df


        total_gene_rows = len(merge.iloc[:,2])
        genes_present = merge.iloc[:,2].count()
        genes_not_in_tso = total_gene_rows - genes_present
        accuracy = round((genes_present/total_gene_rows)*100, 2)


        #get macro data from dfs and combine
        global tso_info
        tso_info = tso.iloc[22:tso_info_end,[0,1]]
        tso_info = tso_info.reset_index(drop=1)
        fm_info = caset.iloc[[0,2]]
        fm_info_i = fm_info.reset_index()
        global fm_info_o
        fm_info_o = fm_info_i.iloc[:,1]
        fm_info_o.loc[2] = ""
        fm_info_o.loc[3] = ""
        fm_info_o.loc[4] = "FM Copy Number Changes"
        fm_info_o = fm_info_o.append(amp_loss_2)
        fm_info_o = fm_info_o.reset_index(drop=1)
        global fm_tso
        fm_tso = pd.concat([fm_info_o, tso_info], axis=1, join="outer")
        
        #to save to current file
        with pd.ExcelWriter(file, engine="openpyxl", mode='a') as writer:
                fm_tso.to_excel(writer, sheet_name="Python_merge", startrow=0, startcol=0)
                merge.to_excel(writer,sheet_name="Python_merge", startrow=27, startcol=0)
                writer.save()

        #formatting
        wb2 = openpyxl.load_workbook(file)
        print(wb2.sheetnames)
        df1 = wb2["Python_merge"]
        df1.delete_cols(1,1)

        bold14Font = Font(size=14, bold=True)
        bold = Font(bold=True)

        border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))


        yellow_highlight = PatternFill(patternType="solid", fgColor=Color('ffff00'))
        green_highlight = PatternFill(patternType="solid", fgColor=Color('66FF00'))

        df1["A27"].fill = yellow_highlight
        df1["C27"].fill = green_highlight
        df1["D27"].fill = green_highlight
        df1["E27"].fill = green_highlight
        df1["F27"].fill = green_highlight
        df1["G27"].fill = green_highlight
        df1["H27"].fill = green_highlight
        df1["I27"].fill = green_highlight

        df1["A27"].font = bold14Font
        df1["C27"].font = bold14Font
        df1["B2"].font = bold
        df1["B7"].font = bold
        df1["B12"].font = bold
        df1["I21"].font = bold
        df1["A6"].font = bold
        df1["A15"].font = bold
        df1["F1"].font = bold14Font
        df1["I21"].border = border
        s = df1["F1:G4"]
        for i in s:
                for c in i:
                        c.border = border

        df1.column_dimensions['A'].width = 35
        df1.column_dimensions['B'].width = 30
        df1.column_dimensions['C'].width = 22
        df1.column_dimensions['D'].width = 13
        df1.column_dimensions['E'].width = 10
        df1.column_dimensions['F'].width = 28
        df1.column_dimensions['G'].width = 25
        df1.column_dimensions['H'].width = 20
        df1.column_dimensions['I'].width = 25
        df1.column_dimensions['J'].width = 15
        df1.column_dimensions['K'].width = 15

        df1["A1"] = "Case Info"
        df1["B1"] = "TSO500 Output"
        df1["C1"] = ""
        df1["A27"] = "Foundation Medicine"
        df1["C27"] = "TSO500"
        df1["D28"] = "Allele Freq"

        df1["A15"] = "MS and TMB:"
        df1["A16"] = str(mstmb)
        df1["F1"] = "Statistics"
        df1["F2"] = "Variants in FM:"
        df1["G2"] = total_gene_rows
        df1["F3"] = "Genes not detected in TSO:"
        df1["G3"] = genes_not_in_tso
        df1["F4"] = "Accuracy:"
        df1["G4"] = accuracy

        wb2.save(file)
